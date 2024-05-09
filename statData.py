from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from statFunctions import *


# def convert_to_png(input_image_path, output_image_path):
#     # Open the image file
#     image = Image.open(input_image_path)
    
#     # Save the image in JPEG format
#     image.save(output_image_path, "PNG") 

images = np.array([['Ressources/GR_300_307-323_fluo_000.tif', 'Ressources/GRMD_307-323_fluo_000.tif', 'Ressources/GRMDT_307-323_fluo_000.tif'],
          ['Ressources/GR_327-353_fluo_000.tif', 'Ressources/GRMD_327-353_fluo_000.tif', 'Ressources/GRMDT_327-353_fluo_000.tif'],
          ['Ressources/GR_370-410_fluo_000.tif', 'Ressources/GRMD_370-410_fluo_000.tif', 'Ressources/GRMDT_370-410_fluo_000.tif'],
          ['Ressources/GR_420-480_fluo_000.tif', 'Ressources/GRMD_420-480_fluo_000.tif', 'Ressources/GRMDT_420-480_fluo_000.tif'],
          ['Ressources/GR_435-455_fluo_000.tif', 'Ressources/GRMD_435-455_fluo_000.tif', 'Ressources/GMDT_435-455_fluo_000.tif']
        ])

images_png = np.array([['Cache/GR_300_307-323_fluo_000.png', 'Cache/GRMD_307-323_fluo_000.png', 'Cache/GRMDT_307-323_fluo_000.png'],
          ['Cache/GR_327-353_fluo_000.png', 'Cache/GRMD_327-353_fluo_000.png', 'Cache/GRMDT_327-353_fluo_000.png'],
          ['Cache/GR_370-410_fluo_000.png', 'Cache/GRMD_370-410_fluo_000.png', 'Cache/GRMDT_370-410_fluo_000.png'],
          ['Cache/GR_420-480_fluo_000.png', 'Cache/GRMD_420-480_fluo_000.png', 'Cache/GRMDT_420-480_fluo_000.png'],
          ['Cache/GR_435-455_fluo_000.png', 'Cache/GRMD_435-455_fluo_000.png', 'Cache/GMDT_435-455_fluo_000.png']
        ])
# for i in range(len(images)):
#     for j in range(len(images[i])):
#         convert_to_png(images[i][j], images_png[i][j])


natures = ['saine', 'malade', 'traite']
statList = ["Autocorrélation", "Energie", "Contraste", "Homogénéité", "Corrélation", "Erreur quadratique moyenne", "SSIM"]
# autocorrelations_matrix = autocorrelations(images)
# energies_matrix = energies(images)
# contrasts_matrix = contrasts(images)
# homogeneities_matrix = homogeneities(images)
# correlations_matrix = correlations(images)
# mse_matrix = mean_squared_errors(images)
# ssim_matrix = ssim(images)

wb = load_workbook('data.xlsx')
selected_paths = {"Image Muscle Sain":"Ressources/GR_327-353_fluo_000.tif",
        "Image Muscle Malade":"Ressources/GRMD_327-353_fluo_000.tif"}
selected_stats = {"Erreur quadratique moyenne":5, "SSIM":6}
def getDataForImage(selected_path, selected_stats):
  data = {}
  for image_type, image_path in selected_path.items():
    image_index = np.where(images == image_path)
    row_index = image_index[0][0] + 1
    col_index = get_column_letter(image_index[1][0] + 1)
    for stat, statIndex in selected_stats.items():
      sheet = wb[stat]
      statValue = float(sheet[col_index + str(row_index)].value)
      data[stat] = statValue
  return data

def getDataForImages(selected_paths, selected_stats):
  data = {}
  images_indexes = []
  for path in selected_paths.values():
    image_row = np.where(images == path)[0][0]
    image_column = np.where(images == path)[1][0]
    images_indexes.append([image_row, image_column])
  
  row_index = (images_indexes[0][0] * 3 + 1) + images_indexes[0][1]
  column_index = get_column_letter(images_indexes[1][1] + 1)
  for stat, statIndex in selected_stats.items():
      sheet = wb[stat]
      statValue = float(sheet[column_index + str(row_index)].value)
      data[stat] = statValue
  return data

